from openpyxl import Workbook, load_workbook
import os
import re
import pandas as pd
import datetime


def out_to_excel(data_list, filename, dirname, title_name):

    os.makedirs('./output/' + dirname, exist_ok=True)
    file_path = './output/' + dirname + '/' + filename

    backup_name = ""

    if os.path.exists(file_path):
        # Excelファイルが存在する場合
        wb = load_workbook(file_path)
        if title_name in wb.sheetnames:
            # 同じシート名が存在する場合
            backup_name = title_name + "_bak"
            wb[title_name].title = backup_name

        wb.create_sheet(title_name)
        target_ws = wb[title_name]

    else:
        # Excelファイルが存在しない場合、Excelワークブックの生成
        wb = Workbook()
        ws = wb.active
        ws.title = title_name
        target_ws = ws

    # ヘッダーを書き込む
    target_ws.append(['掲載カテゴリ', '商品名', '価格(ポイント)', '提供会社'])

    # データを書き込む
    for row in data_list:

        target_ws.append([row['item_category'], row['item_name'],
                          row['item_point'], row['company']])

    if backup_name:
        del wb[backup_name]

    wb.save(file_path)
